# -*- coding:utf-8 _*-
"""
@author:zhouzongquan
@time: 2021/12/7
@email:919839065@qq.com
@function：
"""

import csv
import openpyxl
import time
from itertools import islice
from openpyxl.styles import PatternFill
import contants


def get_data():
    with open(contants.csv_path, 'r',encoding='utf-8') as f:
        reader = csv.reader(f)
        ScripttName = []
        result = []
        for i in islice(reader, 1, None):
            ScripttName.append(i[0])
            result.append(i[2])
        return ScripttName,result


def do_excel():
    workbook = openpyxl.load_workbook(contants.excel_path)
    sheet = workbook['Sheet1']
    max_row = sheet.max_row
    max_column = sheet.max_column
    new_max_column = max_column + 1
    current_time = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
    sheet.cell(1 ,1).value = "脚本名称"
    sheet.cell(1, new_max_column).value = current_time
    ScripttName = get_data()[0]
    result = get_data()[1]
    for s,Sname in enumerate(ScripttName):
        find = False
        for r in range(0,max_row):
            # 读取excel中脚本名称
            ename =sheet.cell(r + 1 ,1).value
            if Sname == ename:
                sheet.cell(r + 1, new_max_column).value = result[s]

                new_result = sheet.cell(r + 1, new_max_column).value
                old_result = sheet.cell(r + 1, new_max_column-1).value
                color = color_value(new_result, old_result)
                sheet.cell(r + 1, new_max_column).fill = PatternFill(fill_type='solid', start_color=color)
                find = True
                break
        if not find:
            max_rows = sheet.max_row
            sheet.cell(max_rows + 1, 1).value = Sname
            sheet.cell(max_rows + 1, new_max_column).value = result[s]

            new_result = sheet.cell(max_rows + 1, new_max_column).value
            old_result = sheet.cell(max_rows + 1, new_max_column-1).value
            color = color_value(new_result, old_result)
            sheet.cell(max_rows + 1, new_max_column).fill = PatternFill(fill_type='solid', start_color=color)
    workbook.save(filename=contants.excel_path)


def color_value(new_result,old_result):
    """
    红色：不通过FF0000
    绿色：通过00FF00
    """
    color = ""
    if new_result == "0" and old_result == "1":
        color = "FF0000"
    elif new_result == "1" and old_result == "0":
        color = "00FF00"
    elif new_result == "0" and old_result == "0":
        color = "000000"
    elif new_result == "1" and old_result == "1":
        color = "000000"
    elif new_result == "1" and old_result == None:
        color = "00FF00"
    elif new_result == "0" and old_result == None:
        color = "FF0000"
    else:
        print("结果有误！")
    return color

#
# def compar():
#     workbook = openpyxl.load_workbook(contants.excel_path)
#     sheet = workbook['Sheet1']
#     max_row = sheet.max_row
#     max_column = sheet.max_column
#     # 循环比对结果
#     for r in range(2,max_row):
#         new_result = sheet.cell(r,max_column).value
#         old_result = sheet.cell(r,max_column-1).value
#         color = color_value(new_result,old_result)
#         sheet.cell(r, max_column).fill = PatternFill(fill_type='solid', start_color=color)
#         workbook.save(filename=contants.excel_path)

if __name__ == '__main__':
    do_excel()

